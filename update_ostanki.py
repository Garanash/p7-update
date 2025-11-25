#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import imaplib
import email
import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import re
import time
import subprocess
import psutil
import requests
import json
import logging
from logging.handlers import RotatingFileHandler
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

EMAIL_LOGIN = "almazgeobur.it@mail.ru"
EMAIL_PASSWORD = "Ba9uV5zDx6rE1fs6PgsV"
IMAP_SERVER = "imap.mail.ru"
IMAP_PORT = 993

MAIN_FILE = "Сборка Москва.xlsx"
TEMP_BOT_FILE = "temp_bot_file.xlsx"
SHEET_OSTANKI = "остатки"
LOG_FILE = "update_ostanki.log"

try:
    from config_p7 import P7_DOC_SERVER_URL, P7_ACCESS_TOKEN, P7_FILE_ID, P7_VERIFY_SSL
except ImportError:
    P7_DOC_SERVER_URL = os.getenv("P7_DOC_SERVER_URL", "")
    P7_ACCESS_TOKEN = os.getenv("P7_ACCESS_TOKEN", "")
    P7_FILE_ID = os.getenv("P7_FILE_ID", "")
    P7_VERIFY_SSL = os.getenv("P7_VERIFY_SSL", "True").lower() == "true"
except AttributeError:
    P7_VERIFY_SSL = False


def setup_logging():
    logger = logging.getLogger('update_ostanki')
    logger.setLevel(logging.DEBUG)
    
    if logger.handlers:
        return logger
    
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    file_handler = RotatingFileHandler(
        LOG_FILE,
        maxBytes=10*1024*1024,
        backupCount=5,
        encoding='utf-8'
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


logger = setup_logging()


def connect_to_email():
    logger.info("Подключение к почтовому серверу...")
    logger.debug(f"Сервер: {IMAP_SERVER}:{IMAP_PORT}, Логин: {EMAIL_LOGIN}")
    try:
        mail = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail.login(EMAIL_LOGIN, EMAIL_PASSWORD)
        mail.select('inbox')
        logger.info("Успешное подключение к почте")
        return mail
    except Exception as e:
        logger.error(f"Ошибка подключения к почте: {e}", exc_info=True)
        return None


def find_latest_excel_attachment(mail):
    logger.info("Поиск Excel файла в письмах...")
    try:
        status, messages = mail.search(None, 'UNSEEN')
        if status != 'OK':
            logger.warning("Не удалось найти письма")
            return None
        
        email_ids = messages[0].split()
        logger.info(f"Найдено непрочитанных писем: {len(email_ids)}")
        
        if not email_ids:
            logger.info("Непрочитанных писем нет, ищем последние 50 писем...")
            status, messages = mail.search(None, 'ALL')
            if status == 'OK':
                all_ids = messages[0].split()
                email_ids = all_ids[-50:] if len(all_ids) > 50 else all_ids
                logger.info(f"Найдено всего писем: {len(all_ids)}, проверяем последние {len(email_ids)}")
        
        excel_files_found = []
        for email_id in reversed(email_ids):
            try:
                status, msg_data = mail.fetch(email_id, '(RFC822)')
                if status != 'OK':
                    continue
                
                email_body = msg_data[0][1]
                email_message = email.message_from_bytes(email_body)
                
                subject = email_message.get('Subject', 'Без темы')
                logger.debug(f"Проверяем письмо: {subject[:50]}...")
                
                for part in email_message.walk():
                    if part.get_content_disposition() == 'attachment':
                        filename = part.get_filename()
                        if filename:
                            try:
                                from email.header import decode_header
                                decoded = decode_header(filename)[0]
                                if decoded[1]:
                                    filename = decoded[0].decode(decoded[1])
                                else:
                                    filename = decoded[0] if isinstance(decoded[0], str) else decoded[0].decode('utf-8')
                            except:
                                pass
                            
                            logger.debug(f"  Найдено вложение: {filename}")
                            if filename and (filename.endswith('.xlsx') or filename.endswith('.xls')):
                                excel_files_found.append((filename, part, email_id))
                                filename_lower = filename.lower()
                                if any(keyword in filename_lower for keyword in ['бот', 'bot', 'xlsx']):
                                    logger.info(f"Найден файл для бота: {filename}")
                                    return part, email_id
            except Exception as e:
                logger.warning(f"Ошибка при обработке письма {email_id}: {e}")
                continue
        
        if excel_files_found:
            filename, part, email_id = excel_files_found[-1]
            logger.info(f"Файл 'для бота' не найден, используем последний Excel файл: {filename}")
            return part, email_id
        
        logger.warning("Excel файлы не найдены в письмах")
        return None
    except Exception as e:
        logger.error(f"Ошибка при поиске вложения: {e}", exc_info=True)
        return None


def download_attachment(part, save_path):
    logger.info(f"Скачивание файла в: {save_path}")
    try:
        with open(save_path, 'wb') as f:
            f.write(part.get_payload(decode=True))
        file_size = os.path.getsize(save_path)
        logger.info(f"Файл успешно скачан: {save_path} (размер: {file_size} байт)")
        return True
    except Exception as e:
        logger.error(f"Ошибка при скачивании файла: {e}", exc_info=True)
        return False


def download_file_from_p7(file_id, save_path):
    try:
        if not P7_DOC_SERVER_URL or P7_DOC_SERVER_URL == "" or "your-p7-doc-server" in P7_DOC_SERVER_URL:
            logger.debug("P7_DOC_SERVER_URL не настроен, пропускаем скачивание файла")
            return False
        
        try:
            verify_ssl = P7_VERIFY_SSL
        except NameError:
            verify_ssl = True
        
        headers = {}
        if P7_ACCESS_TOKEN:
            headers["Authorization"] = f"Bearer {P7_ACCESS_TOKEN}"
        
        base_url = P7_DOC_SERVER_URL.rstrip('/')
        file_id_encoded = requests.utils.quote(str(file_id), safe='')
        
        getfile_paths = [
            f"{base_url}/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/wopi/files/{file_id}/contents",
            f"{base_url}/Products/Files/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/Products/Files/wopi/files/{file_id}/contents",
            f"{base_url}/api/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/api/wopi/files/{file_id}/contents"
        ]
        
        logger.info(f"Попытка скачать файл с Document Server: {file_id}")
        
        for getfile_path in getfile_paths:
            logger.debug(f"Пробуем скачать файл по пути: {getfile_path}")
            try:
                response = requests.get(getfile_path, headers=headers, timeout=30, verify=verify_ssl)
                if response.status_code == 200:
                    with open(save_path, 'wb') as f:
                        f.write(response.content)
                    file_size = os.path.getsize(save_path)
                    logger.info(f"✓ Файл успешно скачан с сервера: {save_path} (размер: {file_size} байт)")
                    return True
                elif response.status_code == 401:
                    logger.warning(f"  Требуется аутентификация для пути: {getfile_path}")
                elif response.status_code == 403:
                    logger.warning(f"  Доступ запрещен для пути: {getfile_path}")
                elif response.status_code != 404:
                    logger.debug(f"  Неожиданный ответ {response.status_code} для пути {getfile_path}")
            except requests.exceptions.RequestException as e:
                logger.debug(f"  Ошибка при скачивании по пути {getfile_path}: {e}")
        
        logger.warning("Не удалось скачать файл с Document Server через WOPI API")
        logger.info("Возможные причины:")
        logger.info("  1. WOPI API не включен на сервере")
        logger.info("  2. Файл не загружен в Document Server через WOPI")
        logger.info("  3. Используется другой API (например, API корпоративного сервера)")
        logger.info("  4. Неправильный file_id или путь к API")
        return False
    except Exception as e:
        logger.error(f"Ошибка при скачивании файла с Document Server: {e}", exc_info=True)
        return False


def update_ostanki_sheet(bot_file_path, main_file_path):
    logger.info(f"Обновление листа '{SHEET_OSTANKI}' из файла: {bot_file_path}")
    try:
        if not os.path.exists(main_file_path):
            logger.warning(f"Файл {main_file_path} не найден локально")
            file_id = P7_FILE_ID if P7_FILE_ID else os.path.basename(main_file_path)
            logger.info(f"Попытка скачать файл с Document Server (ID: {file_id})...")
            if download_file_from_p7(file_id, main_file_path):
                logger.info(f"Файл успешно скачан с сервера")
            else:
                if P7_FILE_ID and P7_FILE_ID != os.path.basename(main_file_path):
                    alt_path = P7_FILE_ID
                    logger.info(f"Пробуем использовать файл с сервера: {alt_path}")
                    if os.path.exists(alt_path):
                        logger.info(f"Найден файл: {alt_path}, используем его")
                        main_file_path = alt_path
                    elif download_file_from_p7(file_id, alt_path):
                        logger.info(f"Файл успешно скачан с сервера как: {alt_path}")
                        main_file_path = alt_path
                    else:
                        logger.error(f"Не удалось скачать файл с сервера и файл отсутствует локально")
                        logger.error(f"Пожалуйста, убедитесь, что файл находится в папке проекта")
                        logger.error(f"Или настройте правильный доступ к Document Server через WOPI API")
                        return None
                else:
                    logger.error(f"Не удалось скачать файл с сервера и файл отсутствует локально")
                    logger.error(f"Пожалуйста, убедитесь, что файл '{main_file_path}' находится в папке проекта")
                    logger.error(f"Или настройте правильный доступ к Document Server через WOPI API")
                    return None
        
        logger.debug("Чтение файла бота...")
        bot_df = pd.read_excel(bot_file_path, sheet_name=0)
        logger.info(f"Прочитано строк из файла бота: {len(bot_df)}")
        logger.debug(f"Столбцы в файле бота: {list(bot_df.columns)}")
        
        for col in bot_df.columns:
            col_str = str(col).strip()
            if 'артикул' in col_str.lower():
                logger.info(f"Нормализация столбца '{col}' (удаление пробелов)")
                bot_df[col] = bot_df[col].astype(str).str.replace(' ', '').str.replace('\t', '').str.replace('\n', '').str.replace('\r', '')
                break
        
        logger.debug(f"Открытие файла: {main_file_path}")
        wb = load_workbook(main_file_path)
        
        existing_sheet = None
        sheets_to_remove = []
        
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower().strip()
            if sheet_lower == SHEET_OSTANKI.lower():
                existing_sheet = sheet_name
            elif sheet_lower.startswith(SHEET_OSTANKI.lower()) and sheet_lower != SHEET_OSTANKI.lower():
                sheets_to_remove.append(sheet_name)
        
        for sheet_name in sheets_to_remove:
            logger.info(f"Удаление дубликата листа: '{sheet_name}'")
            wb.remove(wb[sheet_name])
        
        if existing_sheet:
            ws = wb[existing_sheet]
            if existing_sheet != SHEET_OSTANKI:
                ws.title = SHEET_OSTANKI
            ws.delete_rows(1, ws.max_row)
            logger.info(f"Обновление существующего листа '{SHEET_OSTANKI}'")
        else:
            ws = wb.create_sheet(SHEET_OSTANKI)
            logger.info(f"Создание нового листа '{SHEET_OSTANKI}'")
        
        logger.debug("Запись данных в лист...")
        headers = list(bot_df.columns)
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row_data in enumerate(bot_df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        logger.debug("Сохранение файла...")
        wb.save(main_file_path)
        logger.info(f"Лист '{SHEET_OSTANKI}' успешно обновлен ({len(bot_df)} строк)")
        return bot_df
    except Exception as e:
        logger.error(f"Ошибка при обновлении листа остатки: {e}", exc_info=True)
        return None


def find_catalog_columns(ws):
    catalog_col = None
    catalog_agb_col = None
    ostanki_col = None
    kolvo_col = None
    
    for col in range(1, ws.max_column + 1):
        cell_value = str(ws.cell(row=1, column=col).value or "").lower()
        if 'номер' in cell_value and 'каталог' in cell_value:
            if 'агб' in cell_value:
                catalog_agb_col = col
            else:
                catalog_col = col
        elif 'кол-во' in cell_value and 'станок' in cell_value:
            kolvo_col = col
        elif 'остат' in cell_value and 'склад' in cell_value:
            ostanki_col = col
    
    if kolvo_col and not ostanki_col:
        ostanki_col = kolvo_col + 1
        ws.insert_cols(ostanki_col)
        ws.cell(row=1, column=ostanki_col, value="Остатки на складах")
    
    return catalog_col, catalog_agb_col, ostanki_col


def create_ostanki_dict(ostanki_df):
    logger.info("Создание словаря остатков...")
    ostanki_dict = {}
    
    catalog_cols = []
    ostanki_value_col = None
    
    logger.debug("Поиск столбцов с номерами каталога и остатками...")
    for col in ostanki_df.columns:
        col_lower = str(col).lower().strip()
        col_exact = str(col).strip()
        
        if (col_exact == 'Номер' or 
            col_lower == 'номер' or 
            (col_lower.startswith('номер') and len(col_lower) <= 10 and 'каталог' not in col_lower and 'код' not in col_lower)):
            if col not in catalog_cols:
                catalog_cols.insert(0, col)
        elif 'номенклатура' in col_lower and 'код' in col_lower:
            if col not in catalog_cols:
                catalog_cols.append(col)
        elif 'номенклатура' in col_lower and 'код' not in col_lower:
            if col not in catalog_cols:
                catalog_cols.append(col)
        elif 'остат' in col_lower or ('количество' in col_lower and 'код' not in col_lower) or 'кол-во' in col_lower:
            ostanki_value_col = col
    
    if not catalog_cols:
        logger.error(f"Не найдены столбцы с номерами каталога в листе остатки")
        logger.debug(f"Столбцы: {list(ostanki_df.columns)}")
        return ostanki_dict
    
    if not ostanki_value_col:
        logger.error(f"Не найден столбец с остатками в листе остатки")
        logger.debug(f"Столбцы: {list(ostanki_df.columns)}")
        return ostanki_dict
    
    logger.info(f"Найден столбец с остатками: {ostanki_value_col}")
    
    number_col = None
    for i, col in enumerate(ostanki_df.columns):
        col_str = str(col).strip()
        if (col_str == 'Номер' or 
            col_str.lower().replace(' ', '') == 'номер' or
            (i == 1 and 'номер' in col_str.lower() and 'каталог' not in col_str.lower())):
            number_col = col
            if col not in catalog_cols:
                catalog_cols.insert(0, col)
            break
    
    if number_col:
        logger.info(f"Найден критически важный столбец 'Номер': {number_col}")
    else:
        if len(ostanki_df.columns) > 1:
            potential_num_col = ostanki_df.columns[1]
            logger.info(f"Пробуем использовать столбец по позиции [1]: '{potential_num_col}'")
            if potential_num_col not in catalog_cols:
                catalog_cols.insert(0, potential_num_col)
                number_col = potential_num_col
    
    logger.info(f"Найдены столбцы с номерами каталога (в порядке приоритета): {catalog_cols}")
    
    artikul_col = None
    for col in ostanki_df.columns:
        if 'артикул' in str(col).lower():
            artikul_col = col
            logger.info(f"Найден столбец 'Артикул': {artikul_col}, нормализация...")
            ostanki_df[col] = ostanki_df[col].astype(str).str.replace(' ', '').str.replace('\t', '').str.replace('\n', '').str.replace('\r', '')
            break
    
    logger.debug("Обработка строк для создания словаря остатков...")
    processed_count = 0
    for _, row in ostanki_df.iterrows():
        ostanki_value = row[ostanki_value_col] if pd.notna(row[ostanki_value_col]) else 0
        try:
            ostanki_value = float(ostanki_value)
        except:
            ostanki_value = 0
        
        if ostanki_value == 0:
            continue
        
        catalog_num = None
        for catalog_col in catalog_cols:
            val = row[catalog_col]
            if pd.notna(val):
                catalog_num = str(val).strip()
                catalog_num = catalog_num.replace(' ', '').replace('\t', '').replace('\n', '').replace('\r', '').lower()
                if catalog_num and catalog_num not in ['nan', 'none', '']:
                    if catalog_num in ostanki_dict:
                        ostanki_dict[catalog_num] += ostanki_value
                    else:
                        ostanki_dict[catalog_num] = ostanki_value
                    processed_count += 1
                    break
    
    logger.info(f"Создан словарь остатков для {len(ostanki_dict)} позиций (обработано {processed_count} строк с остатками)")
    return ostanki_dict


def close_file_sessions_p7_api(file_path):
    try:
        file_path_abs = os.path.abspath(file_path)
        file_name = os.path.basename(file_path)
        
        if not P7_DOC_SERVER_URL or P7_DOC_SERVER_URL == "" or "your-p7-doc-server" in P7_DOC_SERVER_URL:
            logger.info("P7_DOC_SERVER_URL не настроен, используем локальное закрытие процессов")
            return close_file_sessions_local(file_path)
        
        logger.info(f"Закрытие сеансов P7-Офис через Document Server API для файла: {file_name}")
        
        try:
            file_id = P7_FILE_ID if P7_FILE_ID else file_name
            
            headers = {}
            if P7_ACCESS_TOKEN:
                headers["Authorization"] = f"Bearer {P7_ACCESS_TOKEN}"
            
            base_url = P7_DOC_SERVER_URL.rstrip('/')
            file_name_encoded = requests.utils.quote(file_name, safe='')
            file_id_encoded = requests.utils.quote(str(file_id), safe='')
            
            wopi_paths = [
                f"{base_url}/Products/Files/wopi/files/{file_id}",
                f"{base_url}/Products/Files/wopi/files/{file_id_encoded}",
                f"{base_url}/wopi/files/{file_id}",
                f"{base_url}/wopi/files/{file_id_encoded}",
                f"{base_url}/Products/api/wopi/files/{file_id}",
                f"{base_url}/Products/api/wopi/files/{file_id_encoded}",
                f"{base_url}/api/wopi/files/{file_id}",
                f"{base_url}/api/wopi/files/{file_id_encoded}",
                f"{base_url}/Products/Files/api/v1/files/{file_id}",
                f"{base_url}/Products/Files/api/v1/files/{file_id_encoded}",
                f"{base_url}/Products/api/v1/files/{file_id}",
                f"{base_url}/Products/api/v1/files/{file_id_encoded}",
                f"{base_url}/api/v1/files/{file_id}",
                f"{base_url}/api/v1/files/{file_id_encoded}",
                f"{base_url}/Products/Files/api/files/{file_id}",
                f"{base_url}/Products/Files/api/files/{file_id_encoded}",
                f"{base_url}/Products/Files/DocEditor.aspx?fileid={file_id}",
                f"{base_url}/wopi/files/{file_name_encoded}",
                f"{base_url}/wopi/files/{file_name}",
                f"{base_url}/Products/Files/wopi/files/{file_name_encoded}",
                f"{base_url}/Products/Files/wopi/files/{file_name}"
            ]
            
            try:
                verify_ssl = P7_VERIFY_SSL
            except NameError:
                verify_ssl = True
            
            wopi_url = None
            response = None
            logger.info(f"Поиск правильного пути к WOPI API для файла ID: {file_id}")
            logger.info(f"Базовый URL: {base_url}")
            
            for wopi_path in wopi_paths:
                test_url = f"{wopi_path}/checkfileinfo"
                logger.info(f"Пробуем путь: {test_url}")
                try:
                    test_response = requests.get(test_url, headers=headers, timeout=10, verify=verify_ssl)
                    logger.info(f"  Ответ: {test_response.status_code} - {test_response.reason}")
                    if test_response.status_code == 200:
                        wopi_url = wopi_path
                        response = test_response
                        logger.info(f"✓ Найден рабочий путь WOPI: {wopi_path}")
                        break
                    elif test_response.status_code == 401:
                        logger.warning(f"  Требуется аутентификация для пути: {wopi_path}")
                    elif test_response.status_code == 403:
                        logger.warning(f"  Доступ запрещен для пути: {wopi_path}")
                    elif test_response.status_code != 404:
                        logger.info(f"  Неожиданный ответ {test_response.status_code} для пути {wopi_path}")
                        if hasattr(test_response, 'text') and test_response.text:
                            logger.debug(f"  Тело ответа: {test_response.text[:200]}")
                except requests.exceptions.RequestException as e:
                    logger.warning(f"  Ошибка при запросе к {test_url}: {e}")
            
            if response and response.status_code == 200:
                file_info = response.json()
                logger.info("Файл найден в Document Server")
                
                if file_info.get("UserCanWrite", False):
                    lock_value = file_info.get("LockValue", "")
                    if lock_value:
                        logger.warning(f"Найдена блокировка файла: {lock_value}")
                        
                        unlock_url = f"{wopi_url}/unlock"
                        unlock_headers = headers.copy()
                        unlock_headers["X-WOPI-Lock"] = lock_value
                        
                        unlock_response = requests.post(unlock_url, headers=unlock_headers, timeout=10, verify=verify_ssl)
                        if unlock_response.status_code == 200:
                            logger.info("Блокировка файла снята через WOPI API")
                        else:
                            logger.warning(f"Не удалось снять блокировку: {unlock_response.status_code}")
                
                base_url = P7_DOC_SERVER_URL.rstrip('/')
                sessions_url = f"{base_url}/api/v1/sessions"
                logger.debug(f"Получение списка сеансов: {sessions_url}")
                sessions_response = requests.get(sessions_url, headers=headers, timeout=10, verify=verify_ssl)
                
                if sessions_response.status_code == 200:
                    sessions = sessions_response.json()
                    file_sessions = [s for s in sessions if file_id in str(s.get("documentId", "")) or file_name in str(s.get("documentName", ""))]
                    
                    if file_sessions:
                        logger.info(f"Найдено {len(file_sessions)} активных сеансов:")
                        for session in file_sessions:
                            session_id = session.get("sessionId", "unknown")
                            user_name = session.get("userName", "unknown")
                            logger.info(f"  - Сеанс {session_id}: пользователь {user_name}")
                        
                        logger.info("Закрытие сеансов...")
                        for session in file_sessions:
                            session_id = session.get("sessionId")
                            if session_id:
                                base_url = P7_DOC_SERVER_URL.rstrip('/')
                                close_paths = [
                                    f"{base_url}/api/v1/sessions/{session_id}",
                                    f"{base_url}/Products/Files/api/v1/sessions/{session_id}",
                                    f"{base_url}/Products/api/v1/sessions/{session_id}",
                                    f"{base_url}/api/sessions/{session_id}"
                                ]
                                
                                closed = False
                                for close_path in close_paths:
                                    logger.debug(f"Пробуем закрыть сеанс по пути: {close_path}")
                                    try:
                                        close_response = requests.delete(close_path, headers=headers, timeout=10, verify=verify_ssl)
                                        if close_response.status_code in [200, 204]:
                                            logger.info(f"  Сеанс {session_id} закрыт")
                                            closed = True
                                            break
                                        elif close_response.status_code != 404:
                                            logger.debug(f"  Ответ {close_response.status_code} для пути {close_path}")
                                    except requests.exceptions.RequestException as e:
                                        logger.debug(f"  Ошибка при закрытии сеанса по пути {close_path}: {e}")
                                
                                if not closed:
                                    logger.warning(f"  Не удалось закрыть сеанс {session_id} ни по одному из путей")
                    else:
                        logger.info("Активных сеансов не найдено")
                else:
                    logger.warning(f"Не удалось получить список сеансов: {sessions_response.status_code}")
            if not response:
                logger.warning("Не удалось найти рабочий путь к WOPI API")
                logger.info("Все пробуемые пути вернули ошибку или 404")
                logger.info("Используем локальное закрытие процессов...")
                return close_file_sessions_local(file_path)
            elif response.status_code == 404:
                logger.warning(f"Файл не найден в Document Server (404)")
                logger.info("Возможные причины:")
                logger.info(f"  1. Файл не загружен в Document Server")
                logger.info(f"  2. Неправильный P7_FILE_ID (текущий: {file_id})")
                logger.info(f"  3. Файл находится по другому пути")
                logger.info("Используем локальное закрытие процессов...")
                return close_file_sessions_local(file_path)
            elif response.status_code != 200:
                logger.warning(f"Ошибка доступа к Document Server: {response.status_code}")
                if hasattr(response, 'text'):
                    logger.debug(f"Ответ сервера: {response.text[:200]}")
                logger.info("Пробуем локальное закрытие процессов...")
                return close_file_sessions_local(file_path)
            
            time.sleep(2)
            
            max_wait = 30
            wait_time = 0
            logger.debug("Ожидание освобождения файла...")
            while wait_time < max_wait:
                try:
                    with open(file_path, 'r+b') as f:
                        pass
                    logger.info("Файл освобожден и готов к обновлению")
                    return True
                except (PermissionError, IOError):
                    wait_time += 1
                    if wait_time % 5 == 0:
                        logger.debug(f"Ожидание освобождения файла... ({wait_time}/{max_wait} сек)")
                    time.sleep(1)
            
            logger.warning(f"Файл не освобожден за {max_wait} секунд, продолжаем...")
            return True
            
        except requests.exceptions.RequestException as e:
            logger.error(f"Ошибка при работе с P7 Document Server API: {e}", exc_info=True)
            logger.info("Пробуем локальное закрытие процессов...")
            return close_file_sessions_local(file_path)
            
    except Exception as e:
        logger.error(f"Ошибка при закрытии сеансов через P7 API: {e}", exc_info=True)
        logger.info("Пробуем локальное закрытие процессов...")
        return close_file_sessions_local(file_path)


def close_file_sessions_local(file_path):
    try:
        file_path_abs = os.path.abspath(file_path)
        logger.info(f"Локальное закрытие процессов для файла: {file_path_abs}")
        
        processes_to_close = []
        logger.debug("Поиск процессов, открывающих файл...")
        for proc in psutil.process_iter(['pid', 'name']):
            try:
                proc_info = proc.info
                proc_name = proc_info['name'].lower()
                
                if 'p7' in proc_name or 'excel' in proc_name or 'spreadsheet' in proc_name or 'calc' in proc_name:
                    try:
                        open_files = proc.open_files()
                        for file_info in open_files:
                            if file_info and hasattr(file_info, 'path'):
                                try:
                                    if os.path.abspath(file_info.path) == file_path_abs:
                                        processes_to_close.append((proc_info['pid'], proc_info['name']))
                                        break
                                except:
                                    pass
                    except (psutil.AccessDenied, psutil.NoSuchProcess):
                        pass
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
                continue
        
        if processes_to_close:
            logger.info(f"Найдено {len(processes_to_close)} процессов, открывающих файл:")
            for pid, name in processes_to_close:
                logger.info(f"  - PID {pid}: {name}")
            
            logger.info("Закрытие процессов...")
            for pid, name in processes_to_close:
                try:
                    proc = psutil.Process(pid)
                    proc.terminate()
                    logger.info(f"  Закрыт процесс {name} (PID {pid})")
                except (psutil.NoSuchProcess, psutil.AccessDenied) as e:
                    logger.warning(f"  Не удалось закрыть процесс {name} (PID {pid}): {e}")
            
            time.sleep(2)
            
            for pid, name in processes_to_close:
                try:
                    proc = psutil.Process(pid)
                    if proc.is_running():
                        proc.kill()
                        logger.info(f"  Принудительно закрыт процесс {name} (PID {pid})")
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            time.sleep(1)
        else:
            logger.info("Файл не открыт в других процессах")
        
        max_wait = 30
        wait_time = 0
        logger.debug("Ожидание освобождения файла...")
        while wait_time < max_wait:
            try:
                with open(file_path, 'r+b') as f:
                    pass
                logger.info("Файл освобожден и готов к обновлению")
                return True
            except (PermissionError, IOError):
                wait_time += 1
                if wait_time % 5 == 0:
                    logger.debug(f"Ожидание освобождения файла... ({wait_time}/{max_wait} сек)")
                time.sleep(1)
        
        logger.warning(f"Файл не освобожден за {max_wait} секунд, продолжаем...")
        return False
    except Exception as e:
        logger.error(f"Ошибка при локальном закрытии сеансов: {e}", exc_info=True)
        return False


def close_file_sessions(file_path):
    return close_file_sessions_p7_api(file_path)


def upload_file_to_p7(file_path):
    try:
        if not P7_DOC_SERVER_URL or P7_DOC_SERVER_URL == "" or "your-p7-doc-server" in P7_DOC_SERVER_URL:
            logger.debug("P7_DOC_SERVER_URL не настроен, пропускаем загрузку файла")
            return True
        
        file_name = os.path.basename(file_path)
        file_id = P7_FILE_ID if P7_FILE_ID else file_name
        
        try:
            verify_ssl = P7_VERIFY_SSL
        except NameError:
            verify_ssl = True
        
        headers = {}
        if P7_ACCESS_TOKEN:
            headers["Authorization"] = f"Bearer {P7_ACCESS_TOKEN}"
        
        base_url = P7_DOC_SERVER_URL.rstrip('/')
        file_id_encoded = requests.utils.quote(str(file_id), safe='')
        file_name_encoded = requests.utils.quote(file_name, safe='')
        
        upload_paths = [
            f"{base_url}/Products/Files/wopi/files/{file_id}/contents",
            f"{base_url}/Products/Files/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/wopi/files/{file_id}/contents",
            f"{base_url}/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/Products/api/wopi/files/{file_id}/contents",
            f"{base_url}/Products/api/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/api/wopi/files/{file_id}/contents",
            f"{base_url}/api/wopi/files/{file_id_encoded}/contents",
            f"{base_url}/Products/Files/api/files/{file_id}/contents",
            f"{base_url}/Products/Files/api/files/{file_id_encoded}/contents",
            f"{base_url}/wopi/files/{file_name_encoded}/contents",
            f"{base_url}/wopi/files/{file_name}/contents",
            f"{base_url}/Products/Files/wopi/files/{file_name_encoded}/contents",
            f"{base_url}/Products/Files/wopi/files/{file_name}/contents"
        ]
        
        logger.info(f"Загрузка обновленного файла в Document Server: {file_name}")
        
        with open(file_path, 'rb') as f:
            file_content = f.read()
        
        headers["X-WOPI-Override"] = "PUT"
        headers["Content-Type"] = "application/octet-stream"
        
        uploaded = False
        response = None
        for upload_path in upload_paths:
            logger.debug(f"Пробуем загрузить файл по пути: {upload_path}")
            try:
                test_response = requests.post(upload_path, headers=headers, data=file_content, timeout=30, verify=verify_ssl)
                logger.debug(f"  Ответ: {test_response.status_code} - {test_response.reason}")
                if test_response.status_code in [200, 201]:
                    logger.info(f"✓ Файл успешно загружен по пути: {upload_path}")
                    uploaded = True
                    response = test_response
                    break
                elif test_response.status_code != 404:
                    logger.debug(f"  Неожиданный ответ {test_response.status_code} для пути {upload_path}")
            except requests.exceptions.RequestException as e:
                logger.debug(f"  Ошибка при загрузке по пути {upload_path}: {e}")
        
        if not uploaded:
            response = type('obj', (object,), {'status_code': 404})()
        
        if response.status_code in [200, 201]:
            logger.info("Файл успешно загружен в Document Server")
            return True
        elif response.status_code == 404:
            logger.info("Файл не найден в Document Server (404) - возможно, файл не загружен в Document Server")
            logger.debug("Продолжаем работу без загрузки файла")
            return False
        else:
            logger.warning(f"Не удалось загрузить файл в Document Server: {response.status_code}")
            logger.debug(f"Ответ сервера: {response.text[:200] if hasattr(response, 'text') else 'N/A'}")
            return False
            
    except Exception as e:
        logger.warning(f"Ошибка при загрузке файла в Document Server: {e}")
        logger.debug("Продолжаем работу без загрузки файла")
        return False


def update_ostanki_in_all_sheets(main_file_path, ostanki_dict):
    logger.info(f"Обновление остатков на всех листах файла: {main_file_path}")
    logger.debug(f"Размер словаря остатков: {len(ostanki_dict)} позиций")
    try:
        wb = load_workbook(main_file_path)
        updated_sheets = []
        total_sheets = len([s for s in wb.sheetnames if s.lower() != SHEET_OSTANKI.lower()])
        logger.info(f"Обработка {total_sheets} листов (кроме '{SHEET_OSTANKI}')...")
        
        for sheet_name in wb.sheetnames:
            if sheet_name.lower() == SHEET_OSTANKI.lower():
                continue
            
            logger.debug(f"Обработка листа: {sheet_name}")
            ws = wb[sheet_name]
            catalog_col, catalog_agb_col, ostanki_col = find_catalog_columns(ws)
            
            if not ostanki_col:
                ostanki_col = ws.max_column + 1
                ws.cell(row=1, column=ostanki_col, value="Остатки на складах")
                logger.debug(f"Создан столбец 'Остатки на складах' в листе '{sheet_name}'")
            
            if catalog_col or catalog_agb_col:
                logger.debug(f"Найдены столбцы каталога: основной={catalog_col}, АГБ={catalog_agb_col}, остатки={ostanki_col}")
                updated_count = 0
                matched_count = 0
                for row in range(2, ws.max_row + 1):
                    catalog_num = None
                    
                    if catalog_col:
                        cell_value = ws.cell(row=row, column=catalog_col).value
                        if cell_value and str(cell_value).strip().lower() != 'none':
                            catalog_num = str(cell_value).strip()
                    
                    if not catalog_num and catalog_agb_col:
                        cell_value = ws.cell(row=row, column=catalog_agb_col).value
                        if cell_value and str(cell_value).strip().lower() != 'none':
                            catalog_num = str(cell_value).strip()
                    
                    if catalog_num and catalog_num.lower() not in ['nan', 'none', '']:
                        catalog_num_clean = str(catalog_num).strip().replace(' ', '').replace('\t', '').replace('\n', '').replace('\r', '').lower()
                        
                        found_value = None
                        if catalog_num_clean in ostanki_dict:
                            found_value = ostanki_dict[catalog_num_clean]
                            matched_count += 1
                        else:
                            for key in ostanki_dict.keys():
                                key_normalized = str(key).replace(' ', '').replace('\t', '').replace('\n', '').replace('\r', '').lower()
                                if catalog_num_clean in key_normalized or key_normalized in catalog_num_clean:
                                    found_value = ostanki_dict[key]
                                    matched_count += 1
                                    break
                        
                        try:
                            cell = ws.cell(row=row, column=ostanki_col)
                            if hasattr(cell, 'value') and hasattr(cell, 'coordinate'):
                                if found_value is not None:
                                    cell.value = found_value
                                    updated_count += 1
                                else:
                                    cell.value = 0
                        except AttributeError:
                            continue
                
                if updated_count > 0:
                    updated_sheets.append(f"{sheet_name} ({updated_count} строк)")
                    logger.info(f"Обновлен лист '{sheet_name}': {updated_count} строк обновлено, {matched_count} совпадений найдено")
                else:
                    logger.debug(f"Лист '{sheet_name}': совпадений не найдено")
            else:
                logger.debug(f"Лист '{sheet_name}': столбцы каталога не найдены, пропуск")
        
        logger.debug("Сохранение файла...")
        wb.save(main_file_path)
        logger.info(f"Всего обновлено листов: {len(updated_sheets)}")
        return updated_sheets
    except Exception as e:
        logger.error(f"Ошибка при обновлении листов: {e}", exc_info=True)
        return []


def main(use_local_file=None):
    logger.info("=" * 50)
    logger.info("Начало обновления остатков")
    logger.info("=" * 50)
    
    start_time = time.time()
    mail = None
    
    try:
        if use_local_file:
            logger.info(f"Использование локального файла: {use_local_file}")
            if not os.path.exists(use_local_file):
                logger.error(f"Файл не найден: {use_local_file}")
                return
            import shutil
            shutil.copy(use_local_file, TEMP_BOT_FILE)
            logger.info(f"Файл скопирован в {TEMP_BOT_FILE}")
        else:
            logger.info("Подключение к почте...")
            mail = connect_to_email()
            if not mail:
                logger.error("Не удалось подключиться к почте")
                return
            
            logger.info("Поиск файла в почте...")
            attachment_data = find_latest_excel_attachment(mail)
            if not attachment_data:
                logger.error("Файл не найден в почте")
                mail.close()
                mail.logout()
                return
            
            part, email_id = attachment_data
            
            logger.info("Скачивание файла...")
            if not download_attachment(part, TEMP_BOT_FILE):
                logger.error("Не удалось скачать файл")
                mail.close()
                mail.logout()
                return
        
        main_file_to_use = MAIN_FILE
        if P7_FILE_ID and P7_FILE_ID.endswith('.xlsx'):
            if os.path.exists(P7_FILE_ID):
                main_file_to_use = P7_FILE_ID
                logger.info(f"Используем файл с сервера: {main_file_to_use}")
            elif not os.path.exists(MAIN_FILE):
                main_file_to_use = P7_FILE_ID
                logger.info(f"Файл {MAIN_FILE} не найден, пробуем использовать файл с сервера: {main_file_to_use}")
        
        logger.info("Закрытие сеансов файла перед обновлением...")
        close_file_sessions(main_file_to_use)
        
        logger.info("Обновление листа 'остатки'...")
        ostanki_df = update_ostanki_sheet(TEMP_BOT_FILE, main_file_to_use)
        if ostanki_df is None:
            logger.error("Не удалось обновить лист остатки")
            if os.path.exists(TEMP_BOT_FILE):
                os.remove(TEMP_BOT_FILE)
            if mail:
                mail.close()
                mail.logout()
            return
        
        logger.info("Создание словаря остатков...")
        ostanki_dict = create_ostanki_dict(ostanki_df)
        
        logger.info("Обновление остатков на всех листах...")
        close_file_sessions(MAIN_FILE)
        updated_sheets = update_ostanki_in_all_sheets(MAIN_FILE, ostanki_dict)
        
        logger.info("Загрузка обновленного файла в Document Server...")
        upload_file_to_p7(MAIN_FILE)
        
        if os.path.exists(TEMP_BOT_FILE):
            os.remove(TEMP_BOT_FILE)
            logger.debug(f"Временный файл удален: {TEMP_BOT_FILE}")
        
        if mail:
            mail.close()
            mail.logout()
            logger.debug("Соединение с почтой закрыто")
        
        elapsed_time = time.time() - start_time
        logger.info("=" * 50)
        logger.info("Обновление завершено успешно!")
        logger.info(f"Время выполнения: {elapsed_time:.2f} секунд")
        logger.info("=" * 50)
        if updated_sheets:
            logger.info("Обновленные листы:")
            for sheet_info in updated_sheets:
                logger.info(f"  - {sheet_info}")
    except Exception as e:
        logger.error(f"Критическая ошибка в main(): {e}", exc_info=True)
        if mail:
            try:
                mail.close()
                mail.logout()
            except:
                pass


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        local_file = sys.argv[1]
        main(use_local_file=local_file)
    else:
        main()

